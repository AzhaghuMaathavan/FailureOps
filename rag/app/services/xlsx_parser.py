import uuid
import logging
from app.models.document import Page, DocumentBlock

logger = logging.getLogger(__name__)

def parse_xlsx_to_blocks(file_path: str, doc_id: str, db):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=True)
    except ImportError:
        logger.warning("openpyxl not installed. Skipping xlsx parsing.")
        return False
    except Exception as e:
        logger.error(f"Failed to load XLSX {file_path}: {e}")
        return False

    for sheet_idx, sheetname in enumerate(wb.sheetnames):
        ws = wb[sheetname]
        
        page_id = str(uuid.uuid4())
        db_page = Page(
            id=page_id,
            document_id=doc_id,
            page_number=sheet_idx,
            image_path="",
            status="COMPLETED"
        )
        db.add(db_page)
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
            
        header = [str(cell) if cell is not None else "" for cell in rows[0]]
        
        block_idx = 0
        chunk_size = 50
        for i in range(1, len(rows), chunk_size):
            chunk_rows = rows[i:i+chunk_size]
            
            lines = [f"| {' | '.join(header)} |"]
            lines.append(f"| {' | '.join(['---'] * len(header))} |")
            
            for row in chunk_rows:
                row_str = [str(cell) if cell is not None else "" for cell in row]
                lines.append(f"| {' | '.join(row_str)} |")
                
            table_md = "\n".join(lines)
            
            db_block = DocumentBlock(
                id=str(uuid.uuid4()),
                document_id=doc_id,
                page_id=page_id,
                block_index=block_idx,
                block_type="Table",
                content=table_md,
                raw_metadata={"sheet": [sheetname], "rows": [f"{i+1}-{min(i+chunk_size, len(rows))}"]}
            )
            db.add(db_block)
            block_idx += 1
            
    db.commit()
    return True
